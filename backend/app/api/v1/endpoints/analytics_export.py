"""
Analytics Export Endpoints for Chrono Scraper FastAPI Application

Comprehensive data export functionality supporting multiple formats:
- JSON (optimized for web applications)
- CSV (spreadsheet compatibility) 
- Parquet (high-performance columnar)
- Excel (business user friendly)
- PDF (report generation)

Features:
- Bulk data export with background processing
- Streaming responses for large datasets
- Compression and optimization
- Export job management and monitoring
- Scheduled report generation
"""

import asyncio
import csv
import io
import json
import logging
import os
import tempfile
import uuid
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from uuid import UUID

# Import pandas with error handling for optional engines
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pd = None
    PANDAS_AVAILABLE = False
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse, FileResponse

# Optional dependencies - handle gracefully if not installed
try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    PARQUET_AVAILABLE = True
except ImportError:
    pa = None
    pq = None
    PARQUET_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = PatternFill = Alignment = None
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    PDF_AVAILABLE = True
except ImportError:
    colors = letter = A4 = getSampleStyleSheet = ParagraphStyle = inch = None
    SimpleDocTemplate = Paragraph = Spacer = Table = TableStyle = None
    PDF_AVAILABLE = False
from sqlmodel import Session
import redis.asyncio as aioredis

from ....core.config import settings
from ....api import deps
from ....models.user import User
from ....services.analytics_service import (
    AnalyticsService, 
    get_analytics_service,
    AnalyticsQueryContext
)
from ....schemas.analytics import (
    AnalyticsExportRequest, AnalyticsExportResponse, AnalyticsExportJob,
    ExportJobStatus, AnalyticsFormat, TimeGranularity,
    BaseAnalyticsResponse, AnalyticsErrorResponse
)

logger = logging.getLogger(__name__)

router = APIRouter()


class ExportJobManager:
    """Manages analytics export jobs with Redis backend"""
    
    def __init__(self):
        self.redis_client: Optional[aioredis.Redis] = None
        self.export_dir = Path(settings.TEMP_DIR) / "analytics_exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
    async def initialize(self):
        """Initialize Redis connection"""
        try:
            self.redis_client = aioredis.from_url(settings.REDIS_URL)
            await self.redis_client.ping()
            logger.info("Export job manager initialized")
        except Exception as e:
            logger.error(f"Failed to initialize export job manager: {e}")
            self.redis_client = None
    
    async def create_job(
        self, 
        user_id: UUID,
        request: AnalyticsExportRequest
    ) -> AnalyticsExportJob:
        """Create new export job"""
        job_id = str(uuid.uuid4())
        
        job = AnalyticsExportJob(
            job_id=job_id,
            status=ExportJobStatus.PENDING,
            created_at=datetime.now()
        )
        
        # Store job in Redis
        if self.redis_client:
            job_data = {
                **job.dict(),
                "user_id": str(user_id),
                "request": request.dict()
            }
            await self.redis_client.setex(
                f"export_job:{job_id}",
                86400,  # 24 hours TTL
                json.dumps(job_data, default=str)
            )
        
        return job
    
    async def get_job(self, job_id: str) -> Optional[AnalyticsExportJob]:
        """Get export job by ID"""
        if not self.redis_client:
            return None
        
        try:
            job_data = await self.redis_client.get(f"export_job:{job_id}")
            if job_data:
                data = json.loads(job_data)
                return AnalyticsExportJob(**data)
        except Exception as e:
            logger.error(f"Error getting job {job_id}: {e}")
        
        return None
    
    async def update_job(
        self, 
        job_id: str, 
        status: ExportJobStatus,
        **kwargs
    ):
        """Update export job status and metadata"""
        if not self.redis_client:
            return
        
        try:
            # Get existing job data
            job_data = await self.redis_client.get(f"export_job:{job_id}")
            if job_data:
                data = json.loads(job_data)
                data["status"] = status.value
                
                # Update specific fields
                for key, value in kwargs.items():
                    if hasattr(AnalyticsExportJob, key):
                        data[key] = value
                
                # Update completion time for completed/failed jobs
                if status in [ExportJobStatus.COMPLETED, ExportJobStatus.FAILED]:
                    data["completed_at"] = datetime.now().isoformat()
                
                # Store updated job
                await self.redis_client.setex(
                    f"export_job:{job_id}",
                    86400,  # 24 hours TTL
                    json.dumps(data, default=str)
                )
        
        except Exception as e:
            logger.error(f"Error updating job {job_id}: {e}")
    
    async def list_user_jobs(self, user_id: UUID) -> List[AnalyticsExportJob]:
        """List export jobs for a user"""
        if not self.redis_client:
            return []
        
        try:
            # This is a simplified implementation
            # In production, you'd want better indexing
            keys = await self.redis_client.keys("export_job:*")
            jobs = []
            
            for key in keys:
                job_data = await self.redis_client.get(key)
                if job_data:
                    data = json.loads(job_data)
                    if data.get("user_id") == str(user_id):
                        jobs.append(AnalyticsExportJob(**data))
            
            # Sort by creation time (newest first)
            jobs.sort(key=lambda x: x.created_at, reverse=True)
            return jobs
            
        except Exception as e:
            logger.error(f"Error listing jobs for user {user_id}: {e}")
            return []
    
    def get_file_path(self, job_id: str, format: AnalyticsFormat) -> Path:
        """Get file path for export job"""
        extensions = {
            AnalyticsFormat.JSON: "json",
            AnalyticsFormat.CSV: "csv",
            AnalyticsFormat.PARQUET: "parquet",
            AnalyticsFormat.EXCEL: "xlsx",
            AnalyticsFormat.PDF: "pdf"
        }
        
        extension = extensions.get(format, "json")
        return self.export_dir / f"{job_id}.{extension}"
    
    async def cleanup_old_files(self, max_age_hours: int = 48):
        """Clean up old export files"""
        try:
            cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
            
            for file_path in self.export_dir.glob("*"):
                if file_path.is_file():
                    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_mtime < cutoff_time:
                        file_path.unlink()
                        logger.info(f"Cleaned up old export file: {file_path}")
                        
        except Exception as e:
            logger.error(f"Error cleaning up export files: {e}")


# Global export job manager
export_manager = ExportJobManager()


class AnalyticsExporter:
    """Handles analytics data export in various formats"""
    
    def __init__(self, analytics_service: AnalyticsService):
        self.analytics_service = analytics_service
    
    async def export_data(
        self,
        job_id: str,
        request: AnalyticsExportRequest,
        user_id: UUID
    ) -> Path:
        """Export analytics data based on request parameters"""
        try:
            # Update job status to processing
            await export_manager.update_job(job_id, ExportJobStatus.PROCESSING)
            
            # Get data based on query type
            data = await self._get_export_data(request, user_id)
            
            # Export to specified format
            file_path = export_manager.get_file_path(job_id, request.format)
            
            if request.format == AnalyticsFormat.JSON:
                await self._export_json(data, file_path, request)
            elif request.format == AnalyticsFormat.CSV:
                await self._export_csv(data, file_path, request)
            elif request.format == AnalyticsFormat.PARQUET:
                await self._export_parquet(data, file_path, request)
            elif request.format == AnalyticsFormat.EXCEL:
                await self._export_excel(data, file_path, request)
            elif request.format == AnalyticsFormat.PDF:
                await self._export_pdf(data, file_path, request)
            else:
                raise ValueError(f"Unsupported export format: {request.format}")
            
            # Update job with completion info
            file_size = file_path.stat().st_size
            download_url = f"/api/v1/analytics/export/download/{job_id}"
            expires_at = datetime.now() + timedelta(hours=48)
            
            await export_manager.update_job(
                job_id,
                ExportJobStatus.COMPLETED,
                file_size=file_size,
                download_url=download_url,
                expires_at=expires_at.isoformat()
            )
            
            return file_path
            
        except Exception as e:
            logger.error(f"Export job {job_id} failed: {e}")
            await export_manager.update_job(
                job_id,
                ExportJobStatus.FAILED,
                error_message=str(e)
            )
            raise
    
    def _validate_export_format(self, format: AnalyticsFormat) -> bool:
        """Validate if export format is supported with current dependencies"""
        if format == AnalyticsFormat.PARQUET and not PARQUET_AVAILABLE:
            return False
        elif format == AnalyticsFormat.EXCEL and not EXCEL_AVAILABLE:
            return False
        elif format == AnalyticsFormat.PDF and not PDF_AVAILABLE:
            return False
        return True
    
    async def _get_export_data(
        self, 
        request: AnalyticsExportRequest,
        user_id: UUID
    ) -> Dict[str, Any]:
        """Get analytics data for export"""
        context = AnalyticsQueryContext(
            user_id=user_id,
            use_cache=False  # Don't use cache for exports
        )
        
        query_type = request.query_type
        params = request.parameters
        
        if query_type == "domain_timeline":
            domain = params.get("domain")
            granularity = TimeGranularity(params.get("granularity", "day"))
            timeline_data = await self.analytics_service.get_domain_timeline(
                domain, granularity, context
            )
            return {
                "query_type": query_type,
                "domain": domain,
                "granularity": granularity.value,
                "data": [point.dict() for point in timeline_data]
            }
        
        elif query_type == "domain_statistics":
            domain = params.get("domain")
            stats = await self.analytics_service.get_domain_statistics(domain, context)
            return {
                "query_type": query_type,
                "domain": domain,
                "data": stats.dict()
            }
        
        elif query_type == "project_performance":
            project_id = UUID(params.get("project_id"))
            performance = await self.analytics_service.get_project_performance(
                project_id, context=context
            )
            return {
                "query_type": query_type,
                "project_id": str(project_id),
                "data": performance.dict()
            }
        
        elif query_type == "top_domains":
            metric = params.get("metric", "total_pages")
            limit = params.get("limit", 100)
            top_domains = await self.analytics_service.get_top_domains(
                metric, limit, context
            )
            return {
                "query_type": query_type,
                "metric": metric,
                "data": [entry.dict() for entry in top_domains]
            }
        
        elif query_type == "system_performance":
            system_data = await self.analytics_service.get_system_performance(
                context=context
            )
            return {
                "query_type": query_type,
                "data": system_data.dict()
            }
        
        else:
            raise ValueError(f"Unsupported query type: {query_type}")
    
    async def _export_json(
        self, 
        data: Dict[str, Any], 
        file_path: Path,
        request: AnalyticsExportRequest
    ):
        """Export data as JSON"""
        export_data = {
            "metadata": {
                "export_type": "analytics",
                "format": "json",
                "exported_at": datetime.now().isoformat(),
                "query_type": data.get("query_type"),
                "include_raw_data": request.include_raw_data
            },
            "data": data if request.include_raw_data else data.get("data", {})
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str, ensure_ascii=False)
    
    async def _export_csv(
        self, 
        data: Dict[str, Any], 
        file_path: Path,
        request: AnalyticsExportRequest
    ):
        """Export data as CSV"""
        query_type = data.get("query_type")
        export_data = data.get("data", {})
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            if query_type == "domain_timeline":
                # Timeline data - each row is a time point
                writer = csv.DictWriter(f, fieldnames=[
                    'timestamp', 'pages_scraped', 'pages_successful', 
                    'pages_failed', 'content_size_mb', 'unique_urls', 'error_rate'
                ])
                writer.writeheader()
                for point in export_data:
                    writer.writerow(point)
            
            elif query_type == "top_domains":
                # Top domains - each row is a domain
                writer = csv.DictWriter(f, fieldnames=[
                    'rank', 'domain', 'total_pages', 'success_rate', 
                    'content_size_mb', 'last_activity', 'projects_count'
                ])
                writer.writeheader()
                for entry in export_data:
                    writer.writerow(entry)
            
            else:
                # Generic key-value export
                writer = csv.writer(f)
                writer.writerow(['Key', 'Value'])
                
                def write_dict(d, prefix=''):
                    for key, value in d.items():
                        full_key = f"{prefix}.{key}" if prefix else key
                        if isinstance(value, dict):
                            write_dict(value, full_key)
                        elif isinstance(value, list):
                            for i, item in enumerate(value):
                                if isinstance(item, dict):
                                    write_dict(item, f"{full_key}[{i}]")
                                else:
                                    writer.writerow([f"{full_key}[{i}]", str(item)])
                        else:
                            writer.writerow([full_key, str(value)])
                
                write_dict(export_data)
    
    async def _export_parquet(
        self, 
        data: Dict[str, Any], 
        file_path: Path,
        request: AnalyticsExportRequest
    ):
        """Export data as Parquet"""
        if not PARQUET_AVAILABLE:
            raise ValueError("Parquet export requires pyarrow package to be installed")
        
        query_type = data.get("query_type")
        export_data = data.get("data", {})
        
        if not PANDAS_AVAILABLE:
            raise ValueError("Parquet export requires pandas package to be installed")
        
        # Convert to pandas DataFrame based on data type
        if query_type == "domain_timeline":
            df = pd.DataFrame(export_data)
            df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        elif query_type == "top_domains":
            df = pd.DataFrame(export_data)
            if 'last_activity' in df.columns:
                df['last_activity'] = pd.to_datetime(df['last_activity'])
        
        else:
            # Flatten nested data for tabular format
            flat_data = self._flatten_dict(export_data)
            df = pd.DataFrame([flat_data])
        
        # Write to Parquet with compression
        try:
            df.to_parquet(
                file_path, 
                compression='gzip',
                index=False,
                engine='pyarrow'
            )
        except ImportError:
            # Fallback to default engine if pyarrow not available
            df.to_parquet(
                file_path, 
                compression='gzip',
                index=False
            )
    
    async def _export_excel(
        self, 
        data: Dict[str, Any], 
        file_path: Path,
        request: AnalyticsExportRequest
    ):
        """Export data as Excel"""
        if not EXCEL_AVAILABLE:
            raise ValueError("Excel export requires openpyxl package to be installed")
        
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        query_type = data.get("query_type")
        export_data = data.get("data", {})
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if query_type == "domain_timeline":
            # Timeline data sheet
            ws = workbook.create_sheet("Domain Timeline")
            headers = [
                'Timestamp', 'Pages Scraped', 'Pages Successful',
                'Pages Failed', 'Content Size (MB)', 'Unique URLs', 'Error Rate (%)'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, point in enumerate(export_data, 2):
                ws.cell(row=row, column=1, value=point.get('timestamp'))
                ws.cell(row=row, column=2, value=point.get('pages_scraped', 0))
                ws.cell(row=row, column=3, value=point.get('pages_successful', 0))
                ws.cell(row=row, column=4, value=point.get('pages_failed', 0))
                ws.cell(row=row, column=5, value=point.get('content_size_mb', 0))
                ws.cell(row=row, column=6, value=point.get('unique_urls', 0))
                ws.cell(row=row, column=7, value=point.get('error_rate', 0))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        elif query_type == "top_domains":
            # Top domains sheet
            ws = workbook.create_sheet("Top Domains")
            headers = [
                'Rank', 'Domain', 'Total Pages', 'Success Rate (%)',
                'Content Size (MB)', 'Last Activity', 'Projects Count'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, entry in enumerate(export_data, 2):
                ws.cell(row=row, column=1, value=entry.get('rank'))
                ws.cell(row=row, column=2, value=entry.get('domain'))
                ws.cell(row=row, column=3, value=entry.get('total_pages'))
                ws.cell(row=row, column=4, value=entry.get('success_rate'))
                ws.cell(row=row, column=5, value=entry.get('content_size_mb'))
                ws.cell(row=row, column=6, value=entry.get('last_activity'))
                ws.cell(row=row, column=7, value=entry.get('projects_count'))
        
        else:
            # Generic data sheet
            ws = workbook.create_sheet("Analytics Data")
            flat_data = self._flatten_dict(export_data)
            
            # Write headers
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, (key, value) in enumerate(flat_data.items(), 2):
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))
        
        # Add metadata sheet
        meta_ws = workbook.create_sheet("Export Metadata", 0)
        metadata = [
            ['Export Type', 'Analytics'],
            ['Format', 'Excel'],
            ['Exported At', datetime.now().isoformat()],
            ['Query Type', query_type],
            ['Include Raw Data', request.include_raw_data]
        ]
        
        for row, (key, value) in enumerate(metadata, 1):
            meta_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            meta_ws.cell(row=row, column=2, value=str(value))
        
        workbook.save(file_path)
    
    async def _export_pdf(
        self, 
        data: Dict[str, Any], 
        file_path: Path,
        request: AnalyticsExportRequest
    ):
        """Export data as PDF report"""
        if not PDF_AVAILABLE:
            raise ValueError("PDF export requires reportlab package to be installed")
        
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        
        query_type = data.get("query_type", "Analytics")
        title = Paragraph(f"Analytics Export Report - {query_type.title()}", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Metadata
        meta_data = [
            ['Export Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ['Query Type', query_type],
            ['Format', 'PDF Report'],
            ['Include Raw Data', str(request.include_raw_data)]
        ]
        
        meta_table = Table(meta_data, colWidths=[2*inch, 3*inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 20))
        
        # Data content
        export_data = data.get("data", {})
        
        if query_type == "domain_statistics":
            # Domain statistics summary
            stats = export_data
            summary_data = [
                ['Total Pages', str(stats.get('total_pages', 0))],
                ['Successful Pages', str(stats.get('successful_pages', 0))],
                ['Failed Pages', str(stats.get('failed_pages', 0))],
                ['Success Rate', f"{stats.get('success_rate', 0):.1f}%"],
                ['Total Content Size', f"{stats.get('total_content_size', 0):.1f} MB"],
                ['Unique URLs', str(stats.get('unique_urls', 0))],
                ['Average Scrape Duration', f"{stats.get('avg_scrape_duration', 0):.2f}s"]
            ]
            
            summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
        
        else:
            # Generic data display
            story.append(Paragraph("Analytics Data:", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Convert data to string representation
            data_text = json.dumps(export_data, indent=2, default=str)
            data_para = Paragraph(f"<pre>{data_text}</pre>", styles['Code'])
            story.append(data_para)
        
        # Build PDF
        doc.build(story)
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
        """Flatten nested dictionary for tabular export"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                for i, item in enumerate(v):
                    if isinstance(item, dict):
                        items.extend(self._flatten_dict(item, f"{new_key}[{i}]", sep=sep).items())
                    else:
                        items.append((f"{new_key}[{i}]", item))
            else:
                items.append((new_key, v))
        return dict(items)


@router.post("/export/bulk-data", response_model=AnalyticsExportResponse)
async def export_analytics_data(
    request: AnalyticsExportRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(deps.get_current_active_user),
    analytics_service: AnalyticsService = Depends(get_analytics_service)
):
    """
    Export analytics data in specified format with background processing
    
    Supports multiple formats and query types with asynchronous processing
    for large datasets.
    """
    try:
        # Validate export format availability
        format_availability = {
            AnalyticsFormat.JSON: True,
            AnalyticsFormat.CSV: True,  
            AnalyticsFormat.PARQUET: PARQUET_AVAILABLE,
            AnalyticsFormat.EXCEL: EXCEL_AVAILABLE,
            AnalyticsFormat.PDF: PDF_AVAILABLE
        }
        
        if not format_availability.get(request.format, False):
            missing_deps = {
                AnalyticsFormat.PARQUET: "pyarrow",
                AnalyticsFormat.EXCEL: "openpyxl", 
                AnalyticsFormat.PDF: "reportlab"
            }
            dep_name = missing_deps.get(request.format, "unknown")
            raise HTTPException(
                status_code=400,
                detail=f"Export format '{request.format}' requires {dep_name} package to be installed"
            )
        
        # Initialize export manager if needed
        if not export_manager.redis_client:
            await export_manager.initialize()
        
        # Create export job
        job = await export_manager.create_job(current_user.id, request)
        
        # Start background export task
        exporter = AnalyticsExporter(analytics_service)
        background_tasks.add_task(
            exporter.export_data,
            job.job_id,
            request,
            current_user.id
        )
        
        # Estimate completion time based on query type and format
        estimated_minutes = {
            "domain_timeline": 2,
            "domain_statistics": 1, 
            "project_performance": 3,
            "top_domains": 2,
            "system_performance": 1
        }.get(request.query_type, 2)
        
        if request.format == AnalyticsFormat.PDF:
            estimated_minutes += 1
        elif request.format == AnalyticsFormat.EXCEL:
            estimated_minutes += 1
        
        estimated_completion = datetime.now() + timedelta(minutes=estimated_minutes)
        
        return AnalyticsExportResponse(
            job=job,
            estimated_completion=estimated_completion
        )
        
    except Exception as e:
        logger.error(f"Export request failed: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/export/jobs", response_model=List[AnalyticsExportJob])
async def list_export_jobs(
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(deps.get_current_active_user)
):
    """List user's export jobs with status and download information"""
    try:
        if not export_manager.redis_client:
            await export_manager.initialize()
        
        jobs = await export_manager.list_user_jobs(current_user.id)
        return jobs[:limit]
        
    except Exception as e:
        logger.error(f"Error listing export jobs: {e}")
        raise HTTPException(status_code=500, detail="Failed to list export jobs")


@router.get("/export/jobs/{job_id}", response_model=AnalyticsExportJob)
async def get_export_job(
    job_id: str,
    current_user: User = Depends(deps.get_current_active_user)
):
    """Get specific export job status and information"""
    try:
        if not export_manager.redis_client:
            await export_manager.initialize()
        
        job = await export_manager.get_job(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Export job not found")
        
        # Verify job belongs to current user
        job_data = await export_manager.redis_client.get(f"export_job:{job_id}")
        if job_data:
            data = json.loads(job_data)
            if data.get("user_id") != str(current_user.id):
                raise HTTPException(status_code=403, detail="Access denied")
        
        return job
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting export job {job_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to get export job")


@router.get("/export/download/{job_id}")
async def download_export_file(
    job_id: str,
    current_user: User = Depends(deps.get_current_active_user)
):
    """Download completed export file"""
    try:
        if not export_manager.redis_client:
            await export_manager.initialize()
        
        # Get job and verify access
        job = await export_manager.get_job(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Export job not found")
        
        # Verify job belongs to current user
        job_data = await export_manager.redis_client.get(f"export_job:{job_id}")
        if job_data:
            data = json.loads(job_data)
            if data.get("user_id") != str(current_user.id):
                raise HTTPException(status_code=403, detail="Access denied")
        else:
            raise HTTPException(status_code=404, detail="Export job not found")
        
        if job.status != ExportJobStatus.COMPLETED:
            raise HTTPException(status_code=400, detail="Export not completed")
        
        # Check if file has expired
        if job.expires_at and datetime.fromisoformat(job.expires_at) < datetime.now():
            raise HTTPException(status_code=410, detail="Download link expired")
        
        # Find the file (check all possible formats)
        file_path = None
        for format in AnalyticsFormat:
            potential_path = export_manager.get_file_path(job_id, format)
            if potential_path.exists():
                file_path = potential_path
                break
        
        if not file_path or not file_path.exists():
            raise HTTPException(status_code=404, detail="Export file not found")
        
        # Determine media type based on file extension
        media_types = {
            '.json': 'application/json',
            '.csv': 'text/csv',
            '.parquet': 'application/parquet',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.pdf': 'application/pdf'
        }
        
        media_type = media_types.get(file_path.suffix, 'application/octet-stream')
        
        return FileResponse(
            path=str(file_path),
            media_type=media_type,
            filename=f"analytics_export_{job_id}{file_path.suffix}",
            headers={"Content-Disposition": f"attachment; filename=analytics_export_{job_id}{file_path.suffix}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading export file {job_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to download export file")


@router.delete("/export/jobs/{job_id}")
async def delete_export_job(
    job_id: str,
    current_user: User = Depends(deps.get_current_active_user)
):
    """Delete export job and associated files"""
    try:
        if not export_manager.redis_client:
            await export_manager.initialize()
        
        # Verify job access
        job_data = await export_manager.redis_client.get(f"export_job:{job_id}")
        if job_data:
            data = json.loads(job_data)
            if data.get("user_id") != str(current_user.id):
                raise HTTPException(status_code=403, detail="Access denied")
        else:
            raise HTTPException(status_code=404, detail="Export job not found")
        
        # Delete Redis job data
        await export_manager.redis_client.delete(f"export_job:{job_id}")
        
        # Delete associated files
        for format in AnalyticsFormat:
            file_path = export_manager.get_file_path(job_id, format)
            if file_path.exists():
                file_path.unlink()
        
        return {"message": "Export job deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting export job {job_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete export job")


@router.post("/export/cleanup")
async def cleanup_old_exports(
    max_age_hours: int = Query(48, ge=1, le=168),
    current_user: User = Depends(deps.get_current_admin_user)
):
    """Clean up old export files (admin only)"""
    try:
        if not export_manager.redis_client:
            await export_manager.initialize()
        
        await export_manager.cleanup_old_files(max_age_hours)
        
        return BaseAnalyticsResponse(
            data={"message": f"Cleaned up export files older than {max_age_hours} hours"},
            metadata={"max_age_hours": max_age_hours}
        )
        
    except Exception as e:
        logger.error(f"Error cleaning up exports: {e}")
        raise HTTPException(status_code=500, detail="Failed to clean up exports")