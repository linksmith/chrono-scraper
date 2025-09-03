"""
Advanced compliance reporting and export system for GDPR, SOX, HIPAA, and other regulatory frameworks
"""
import csv
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass, asdict
from enum import Enum
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from jinja2 import Template

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import and_
from sqlmodel import select

from app.models.audit_log import (
    AuditLog, 
    AuditCategory, 
    SeverityLevel, 
    AuditActions
)


logger = logging.getLogger(__name__)


class ComplianceFramework(str, Enum):
    """Supported compliance frameworks"""
    GDPR = "gdpr"           # General Data Protection Regulation
    SOX = "sox"             # Sarbanes-Oxley Act
    HIPAA = "hipaa"         # Health Insurance Portability and Accountability Act
    PCI_DSS = "pci_dss"     # Payment Card Industry Data Security Standard
    ISO_27001 = "iso_27001" # Information Security Management
    NIST = "nist"           # National Institute of Standards and Technology
    CUSTOM = "custom"       # Custom compliance requirements


class ReportFormat(str, Enum):
    """Supported report formats"""
    JSON = "json"
    CSV = "csv"
    PDF = "pdf"
    HTML = "html"
    XML = "xml"
    XLSX = "xlsx"


@dataclass
class ComplianceRequirement:
    """Individual compliance requirement definition"""
    requirement_id: str
    framework: ComplianceFramework
    title: str
    description: str
    mandatory_fields: List[str]
    audit_categories: List[AuditCategory]
    audit_actions: List[str]
    retention_period_days: int
    severity_threshold: SeverityLevel
    automated_check: bool = True
    
    
@dataclass
class ComplianceViolation:
    """Detected compliance violation"""
    violation_id: str
    requirement_id: str
    framework: ComplianceFramework
    severity: str  # low, medium, high, critical
    title: str
    description: str
    detected_at: datetime
    affected_records: List[int]  # audit log IDs
    evidence: Dict[str, Any]
    remediation_steps: List[str]
    status: str = "open"  # open, investigating, resolved, false_positive


@dataclass
class ComplianceReport:
    """Complete compliance report"""
    report_id: str
    framework: ComplianceFramework
    generated_at: datetime
    report_period_start: datetime
    report_period_end: datetime
    generated_by: Optional[int]  # admin user ID
    
    # Executive Summary
    executive_summary: Dict[str, Any]
    
    # Compliance Status
    total_requirements: int
    compliant_requirements: int
    violation_count: int
    compliance_score: float  # 0-100
    
    # Detailed Results
    requirements_status: List[Dict[str, Any]]
    violations: List[ComplianceViolation]
    recommendations: List[str]
    
    # Supporting Data
    audit_log_summary: Dict[str, Any]
    data_processing_activities: List[Dict[str, Any]]
    risk_assessment: Dict[str, Any]
    
    # Metadata
    report_metadata: Dict[str, Any]


class ComplianceReportingService:
    """
    Advanced compliance reporting service providing:
    - Multi-framework compliance assessment
    - Automated violation detection
    - Executive reporting with risk scoring
    - Detailed audit trail analysis
    - Export to multiple formats
    - Scheduled compliance monitoring
    - Evidence collection and documentation
    - Remediation tracking
    """
    
    def __init__(self):
        self.compliance_requirements = self._init_compliance_requirements()
        self.report_templates = self._init_report_templates()
        
    def _init_compliance_requirements(self) -> Dict[ComplianceFramework, List[ComplianceRequirement]]:
        """Initialize compliance requirements for each framework"""
        return {
            ComplianceFramework.GDPR: [
                ComplianceRequirement(
                    requirement_id="GDPR-ART-30",
                    framework=ComplianceFramework.GDPR,
                    title="Records of Processing Activities",
                    description="Maintain records of all personal data processing activities",
                    mandatory_fields=['user_id', 'action', 'created_at', 'ip_address'],
                    audit_categories=[AuditCategory.USER_MANAGEMENT, AuditCategory.COMPLIANCE],
                    audit_actions=[
                        AuditActions.USER_CREATE,
                        AuditActions.USER_UPDATE,
                        AuditActions.USER_DELETE,
                        AuditActions.GDPR_REQUEST,
                        AuditActions.GDPR_DATA_EXPORT,
                        AuditActions.GDPR_DATA_DELETION
                    ],
                    retention_period_days=1095,  # 3 years
                    severity_threshold=SeverityLevel.MEDIUM
                ),
                
                ComplianceRequirement(
                    requirement_id="GDPR-ART-32",
                    framework=ComplianceFramework.GDPR,
                    title="Security of Processing",
                    description="Implement appropriate technical and organizational measures",
                    mandatory_fields=['user_id', 'action', 'success', 'ip_address'],
                    audit_categories=[AuditCategory.SECURITY_EVENT, AuditCategory.AUTHENTICATION],
                    audit_actions=[
                        AuditActions.USER_LOGIN_FAILED,
                        AuditActions.UNAUTHORIZED_ACCESS,
                        AuditActions.SECURITY_VULNERABILITY_DETECTED
                    ],
                    retention_period_days=2190,  # 6 years
                    severity_threshold=SeverityLevel.HIGH
                ),
                
                ComplianceRequirement(
                    requirement_id="GDPR-ART-33",
                    framework=ComplianceFramework.GDPR,
                    title="Notification of Personal Data Breach",
                    description="Document and report personal data breaches",
                    mandatory_fields=['user_id', 'action', 'severity', 'details'],
                    audit_categories=[AuditCategory.SECURITY_EVENT],
                    audit_actions=[
                        AuditActions.BREACH_ATTEMPT,
                        AuditActions.DATA_BREACH_ATTEMPT,
                        AuditActions.UNAUTHORIZED_ACCESS
                    ],
                    retention_period_days=2190,  # 6 years
                    severity_threshold=SeverityLevel.HIGH
                )
            ],
            
            ComplianceFramework.SOX: [
                ComplianceRequirement(
                    requirement_id="SOX-302",
                    framework=ComplianceFramework.SOX,
                    title="Corporate Responsibility for Financial Reports",
                    description="Maintain controls over financial reporting systems",
                    mandatory_fields=['admin_user_id', 'action', 'created_at', 'success'],
                    audit_categories=[AuditCategory.SYSTEM_CONFIG, AuditCategory.USER_MANAGEMENT],
                    audit_actions=[
                        AuditActions.SYSTEM_CONFIG_UPDATE,
                        AuditActions.USER_ROLE_ASSIGN,
                        AuditActions.USER_PERMISSION_GRANT,
                        AuditActions.ADMIN_LOGIN
                    ],
                    retention_period_days=2555,  # 7 years
                    severity_threshold=SeverityLevel.HIGH
                ),
                
                ComplianceRequirement(
                    requirement_id="SOX-404",
                    framework=ComplianceFramework.SOX,
                    title="Management Assessment of Internal Controls",
                    description="Document and assess internal control effectiveness",
                    mandatory_fields=['admin_user_id', 'action', 'before_values', 'after_values'],
                    audit_categories=[AuditCategory.SYSTEM_CONFIG, AuditCategory.USER_MANAGEMENT],
                    audit_actions=[
                        AuditActions.SYSTEM_CONFIG_UPDATE,
                        AuditActions.USER_ROLE_ASSIGN,
                        AuditActions.USER_PERMISSION_GRANT,
                        AuditActions.BULK_USER_ROLE_ASSIGN
                    ],
                    retention_period_days=2555,  # 7 years
                    severity_threshold=SeverityLevel.MEDIUM
                )
            ],
            
            ComplianceFramework.HIPAA: [
                ComplianceRequirement(
                    requirement_id="HIPAA-164.312",
                    framework=ComplianceFramework.HIPAA,
                    title="Technical Safeguards",
                    description="Implement technical safeguards for PHI",
                    mandatory_fields=['user_id', 'action', 'ip_address', 'success'],
                    audit_categories=[AuditCategory.AUTHENTICATION, AuditCategory.SECURITY_EVENT],
                    audit_actions=[
                        AuditActions.USER_LOGIN,
                        AuditActions.USER_LOGIN_FAILED,
                        AuditActions.USER_PROFILE_VIEW,
                        AuditActions.DATA_EXPORT
                    ],
                    retention_period_days=2190,  # 6 years
                    severity_threshold=SeverityLevel.MEDIUM
                ),
                
                ComplianceRequirement(
                    requirement_id="HIPAA-164.308",
                    framework=ComplianceFramework.HIPAA,
                    title="Administrative Safeguards",
                    description="Implement administrative safeguards for PHI",
                    mandatory_fields=['admin_user_id', 'action', 'resource_type'],
                    audit_categories=[AuditCategory.USER_MANAGEMENT, AuditCategory.SYSTEM_CONFIG],
                    audit_actions=[
                        AuditActions.USER_ROLE_ASSIGN,
                        AuditActions.USER_PERMISSION_GRANT,
                        AuditActions.SYSTEM_CONFIG_UPDATE
                    ],
                    retention_period_days=2190,  # 6 years
                    severity_threshold=SeverityLevel.HIGH
                )
            ]
        }
    
    def _init_report_templates(self) -> Dict[str, str]:
        """Initialize report templates for different formats"""
        return {
            'html_executive_summary': '''
            <div class="executive-summary">
                <h2>Executive Summary</h2>
                <div class="compliance-score">
                    <h3>Overall Compliance Score: {{ compliance_score }}%</h3>
                    <p>Compliance Level: 
                        {% if compliance_score >= 95 %}
                            <span class="excellent">Excellent</span>
                        {% elif compliance_score >= 80 %}
                            <span class="good">Good</span>
                        {% elif compliance_score >= 60 %}
                            <span class="fair">Fair</span>
                        {% else %}
                            <span class="poor">Needs Improvement</span>
                        {% endif %}
                    </p>
                </div>
                
                <div class="key-metrics">
                    <h3>Key Metrics</h3>
                    <ul>
                        <li>Total Requirements Assessed: {{ total_requirements }}</li>
                        <li>Compliant Requirements: {{ compliant_requirements }}</li>
                        <li>Violations Detected: {{ violation_count }}</li>
                        <li>High/Critical Violations: {{ high_critical_violations }}</li>
                    </ul>
                </div>
                
                {% if recommendations %}
                <div class="recommendations">
                    <h3>Key Recommendations</h3>
                    <ol>
                        {% for recommendation in recommendations[:5] %}
                        <li>{{ recommendation }}</li>
                        {% endfor %}
                    </ol>
                </div>
                {% endif %}
            </div>
            ''',
            
            'html_violations': '''
            <div class="violations-section">
                <h2>Compliance Violations</h2>
                {% if violations %}
                    <table class="violations-table">
                        <thead>
                            <tr>
                                <th>Severity</th>
                                <th>Requirement</th>
                                <th>Description</th>
                                <th>Detected</th>
                                <th>Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for violation in violations %}
                            <tr class="severity-{{ violation.severity }}">
                                <td>{{ violation.severity|title }}</td>
                                <td>{{ violation.requirement_id }}</td>
                                <td>{{ violation.title }}</td>
                                <td>{{ violation.detected_at.strftime('%Y-%m-%d %H:%M') }}</td>
                                <td>{{ violation.status|title }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                {% else %}
                    <p class="no-violations">No compliance violations detected in this period.</p>
                {% endif %}
            </div>
            '''
        }
    
    async def generate_compliance_report(
        self,
        framework: ComplianceFramework,
        start_date: datetime,
        end_date: datetime,
        generated_by: Optional[int],
        db: AsyncSession
    ) -> ComplianceReport:
        """Generate comprehensive compliance report for specified framework"""
        
        report_id = f"{framework.value}_{int(datetime.now(timezone.utc).timestamp())}"
        
        # Get compliance requirements for framework
        requirements = self.compliance_requirements.get(framework, [])
        
        # Assess each requirement
        requirements_status = []
        all_violations = []
        
        for requirement in requirements:
            requirement_result = await self._assess_compliance_requirement(
                requirement, start_date, end_date, db
            )
            
            requirements_status.append(requirement_result['status'])
            all_violations.extend(requirement_result['violations'])
        
        # Calculate compliance metrics
        compliant_count = sum(1 for status in requirements_status if status['compliant'])
        compliance_score = (compliant_count / len(requirements) * 100) if requirements else 100
        
        # Generate audit log summary
        audit_summary = await self._generate_audit_summary(
            framework, start_date, end_date, db
        )
        
        # Generate data processing activities summary
        processing_activities = await self._generate_processing_activities(
            framework, start_date, end_date, db
        )
        
        # Perform risk assessment
        risk_assessment = await self._perform_risk_assessment(
            all_violations, requirements_status
        )
        
        # Generate recommendations
        recommendations = await self._generate_recommendations(
            all_violations, requirements_status, compliance_score
        )
        
        # Create executive summary
        executive_summary = {
            'framework': framework.value.upper(),
            'assessment_period': f"{start_date.date()} to {end_date.date()}",
            'compliance_score': round(compliance_score, 2),
            'total_audit_events': audit_summary.get('total_events', 0),
            'security_incidents': audit_summary.get('security_incidents', 0),
            'risk_level': risk_assessment.get('overall_risk_level', 'medium'),
            'key_findings': risk_assessment.get('key_findings', [])
        }
        
        return ComplianceReport(
            report_id=report_id,
            framework=framework,
            generated_at=datetime.now(timezone.utc),
            report_period_start=start_date,
            report_period_end=end_date,
            generated_by=generated_by,
            executive_summary=executive_summary,
            total_requirements=len(requirements),
            compliant_requirements=compliant_count,
            violation_count=len(all_violations),
            compliance_score=compliance_score,
            requirements_status=requirements_status,
            violations=all_violations,
            recommendations=recommendations,
            audit_log_summary=audit_summary,
            data_processing_activities=processing_activities,
            risk_assessment=risk_assessment,
            report_metadata={
                'generation_duration_ms': 0,  # Would be calculated
                'data_sources': ['audit_logs'],
                'report_version': '1.0',
                'compliance_framework_version': '2023'
            }
        )
    
    async def _assess_compliance_requirement(
        self,
        requirement: ComplianceRequirement,
        start_date: datetime,
        end_date: datetime,
        db: AsyncSession
    ) -> Dict[str, Any]:
        """Assess a specific compliance requirement"""
        
        violations = []
        
        # Get relevant audit logs
        query = select(AuditLog).where(
            and_(
                AuditLog.created_at >= start_date,
                AuditLog.created_at <= end_date
            )
        )
        
        # Apply requirement filters
        if requirement.audit_categories:
            query = query.where(AuditLog.category.in_(requirement.audit_categories))
        
        if requirement.audit_actions:
            query = query.where(AuditLog.action.in_(requirement.audit_actions))
        
        result = await db.execute(query)
        relevant_logs = result.scalars().all()
        
        # Check for violations
        violations.extend(await self._check_data_completeness(requirement, relevant_logs))
        violations.extend(await self._check_security_incidents(requirement, relevant_logs))
        violations.extend(await self._check_access_controls(requirement, relevant_logs))
        violations.extend(await self._check_retention_compliance(requirement, relevant_logs, db))
        
        # Determine compliance status
        is_compliant = len(violations) == 0
        confidence_score = self._calculate_confidence_score(requirement, relevant_logs, violations)
        
        return {
            'status': {
                'requirement_id': requirement.requirement_id,
                'title': requirement.title,
                'compliant': is_compliant,
                'confidence_score': confidence_score,
                'assessed_records': len(relevant_logs),
                'violations_found': len(violations),
                'last_assessment': datetime.now(timezone.utc).isoformat()
            },
            'violations': violations
        }
    
    async def _check_data_completeness(
        self,
        requirement: ComplianceRequirement,
        audit_logs: List[AuditLog]
    ) -> List[ComplianceViolation]:
        """Check for data completeness violations"""
        violations = []
        
        for log in audit_logs:
            missing_fields = []
            
            for field in requirement.mandatory_fields:
                if not getattr(log, field, None):
                    missing_fields.append(field)
            
            if missing_fields:
                violation = ComplianceViolation(
                    violation_id=f"{requirement.requirement_id}_incomplete_{log.id}",
                    requirement_id=requirement.requirement_id,
                    framework=requirement.framework,
                    severity="medium",
                    title="Incomplete Audit Record",
                    description=f"Audit log {log.id} missing required fields: {', '.join(missing_fields)}",
                    detected_at=datetime.now(timezone.utc),
                    affected_records=[log.id],
                    evidence={'missing_fields': missing_fields, 'log_action': log.action},
                    remediation_steps=[
                        "Review audit logging configuration",
                        "Ensure all mandatory fields are captured",
                        "Update data collection procedures"
                    ]
                )
                violations.append(violation)
        
        return violations
    
    async def _check_security_incidents(
        self,
        requirement: ComplianceRequirement,
        audit_logs: List[AuditLog]
    ) -> List[ComplianceViolation]:
        """Check for security incident violations"""
        violations = []
        
        # Look for high-severity security events
        security_events = [
            log for log in audit_logs
            if (log.category == AuditCategory.SECURITY_EVENT and 
                log.severity in [SeverityLevel.HIGH, SeverityLevel.CRITICAL])
        ]
        
        for event in security_events:
            violation = ComplianceViolation(
                violation_id=f"{requirement.requirement_id}_security_{event.id}",
                requirement_id=requirement.requirement_id,
                framework=requirement.framework,
                severity="high" if event.severity == SeverityLevel.HIGH else "critical",
                title="Security Incident Detected",
                description=f"High-severity security event: {event.action}",
                detected_at=event.created_at,
                affected_records=[event.id],
                evidence={
                    'event_action': event.action,
                    'severity': event.severity,
                    'ip_address': event.ip_address,
                    'user_id': event.user_id
                },
                remediation_steps=[
                    "Investigate security incident",
                    "Implement corrective measures",
                    "Review security controls",
                    "Document incident response"
                ]
            )
            violations.append(violation)
        
        return violations
    
    async def _check_access_controls(
        self,
        requirement: ComplianceRequirement,
        audit_logs: List[AuditLog]
    ) -> List[ComplianceViolation]:
        """Check for access control violations"""
        violations = []
        
        # Check for failed authentication attempts
        failed_logins = [
            log for log in audit_logs
            if log.action == AuditActions.USER_LOGIN_FAILED
        ]
        
        # Group by IP and check for brute force patterns
        from collections import defaultdict
        ip_failures = defaultdict(list)
        
        for log in failed_logins:
            if log.ip_address:
                ip_failures[log.ip_address].append(log)
        
        # Check for suspicious patterns
        for ip, failures in ip_failures.items():
            if len(failures) > 10:  # More than 10 failures from same IP
                violation = ComplianceViolation(
                    violation_id=f"{requirement.requirement_id}_bruteforce_{ip}",
                    requirement_id=requirement.requirement_id,
                    framework=requirement.framework,
                    severity="high",
                    title="Potential Brute Force Attack",
                    description=f"Multiple failed login attempts from IP {ip}",
                    detected_at=max(f.created_at for f in failures),
                    affected_records=[f.id for f in failures],
                    evidence={
                        'ip_address': ip,
                        'failure_count': len(failures),
                        'time_span_hours': (max(f.created_at for f in failures) - 
                                          min(f.created_at for f in failures)).total_seconds() / 3600
                    },
                    remediation_steps=[
                        f"Block IP address {ip}",
                        "Implement rate limiting",
                        "Review authentication logs",
                        "Consider implementing CAPTCHA"
                    ]
                )
                violations.append(violation)
        
        return violations
    
    async def _check_retention_compliance(
        self,
        requirement: ComplianceRequirement,
        audit_logs: List[AuditLog],
        db: AsyncSession
    ) -> List[ComplianceViolation]:
        """Check for data retention compliance violations"""
        violations = []
        
        # Check for logs older than retention period that should have been archived
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=requirement.retention_period_days)
        
        old_logs_query = select(AuditLog).where(
            and_(
                AuditLog.created_at < cutoff_date,
                AuditLog.archived is False
            )
        )
        
        if requirement.audit_categories:
            old_logs_query = old_logs_query.where(AuditLog.category.in_(requirement.audit_categories))
        
        result = await db.execute(old_logs_query)
        unarchived_old_logs = result.scalars().all()
        
        if unarchived_old_logs:
            violation = ComplianceViolation(
                violation_id=f"{requirement.requirement_id}_retention_violation",
                requirement_id=requirement.requirement_id,
                framework=requirement.framework,
                severity="medium",
                title="Data Retention Policy Violation",
                description=f"Found {len(unarchived_old_logs)} audit logs older than {requirement.retention_period_days} days that are not archived",
                detected_at=datetime.now(timezone.utc),
                affected_records=[log.id for log in unarchived_old_logs],
                evidence={
                    'retention_period_days': requirement.retention_period_days,
                    'unarchived_count': len(unarchived_old_logs),
                    'oldest_log_date': min(log.created_at for log in unarchived_old_logs).isoformat()
                },
                remediation_steps=[
                    "Archive old audit logs",
                    "Review retention policy implementation",
                    "Automate archival process",
                    "Update data lifecycle management"
                ]
            )
            violations.append(violation)
        
        return violations
    
    def _calculate_confidence_score(
        self,
        requirement: ComplianceRequirement,
        audit_logs: List[AuditLog],
        violations: List[ComplianceViolation]
    ) -> float:
        """Calculate confidence score for compliance assessment"""
        
        if not audit_logs:
            return 0.0
        
        # Base score from data completeness
        complete_records = 0
        for log in audit_logs:
            missing_fields = sum(1 for field in requirement.mandatory_fields 
                               if not getattr(log, field, None))
            if missing_fields == 0:
                complete_records += 1
        
        completeness_score = complete_records / len(audit_logs)
        
        # Penalty for violations
        violation_penalty = min(len(violations) * 0.1, 0.5)  # Max 50% penalty
        
        # Final confidence score
        confidence = max(0.0, completeness_score - violation_penalty)
        
        return round(confidence * 100, 2)
    
    async def export_compliance_report(
        self,
        report: ComplianceReport,
        format: ReportFormat,
        include_evidence: bool = False
    ) -> Union[str, bytes]:
        """Export compliance report in specified format"""
        
        if format == ReportFormat.JSON:
            return await self._export_json(report)
        elif format == ReportFormat.CSV:
            return await self._export_csv(report)
        elif format == ReportFormat.PDF:
            return await self._export_pdf(report)
        elif format == ReportFormat.HTML:
            return await self._export_html(report)
        elif format == ReportFormat.XML:
            return await self._export_xml(report)
        elif format == ReportFormat.XLSX:
            return await self._export_xlsx(report)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def _export_json(self, report: ComplianceReport) -> str:
        """Export report as JSON"""
        report_dict = asdict(report)
        
        # Convert datetime objects to ISO format
        def convert_datetime(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            return obj
        
        return json.dumps(report_dict, default=convert_datetime, indent=2)
    
    async def _export_csv(self, report: ComplianceReport) -> str:
        """Export report as CSV"""
        output = StringIO()
        
        # Write executive summary
        writer = csv.writer(output)
        writer.writerow(['Compliance Report - Executive Summary'])
        writer.writerow(['Framework', report.framework.value.upper()])
        writer.writerow(['Generated At', report.generated_at.isoformat()])
        writer.writerow(['Period', f"{report.report_period_start.date()} to {report.report_period_end.date()}"])
        writer.writerow(['Compliance Score', f"{report.compliance_score}%"])
        writer.writerow(['Total Requirements', report.total_requirements])
        writer.writerow(['Compliant Requirements', report.compliant_requirements])
        writer.writerow(['Violations', report.violation_count])
        writer.writerow([])
        
        # Write requirements status
        writer.writerow(['Requirements Assessment'])
        writer.writerow(['Requirement ID', 'Title', 'Compliant', 'Confidence Score', 'Violations'])
        
        for req_status in report.requirements_status:
            writer.writerow([
                req_status['requirement_id'],
                req_status['title'],
                'Yes' if req_status['compliant'] else 'No',
                f"{req_status['confidence_score']}%",
                req_status['violations_found']
            ])
        
        writer.writerow([])
        
        # Write violations
        if report.violations:
            writer.writerow(['Violations'])
            writer.writerow(['Violation ID', 'Severity', 'Title', 'Description', 'Detected At', 'Status'])
            
            for violation in report.violations:
                writer.writerow([
                    violation.violation_id,
                    violation.severity,
                    violation.title,
                    violation.description,
                    violation.detected_at.isoformat(),
                    violation.status
                ])
        
        return output.getvalue()
    
    async def _export_pdf(self, report: ComplianceReport) -> bytes:
        """Export report as PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        
        story.append(Paragraph(f"{report.framework.value.upper()} Compliance Report", title_style))
        story.append(Spacer(1, 12))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Assessment Period', f"{report.report_period_start.date()} to {report.report_period_end.date()}"],
            ['Compliance Score', f"{report.compliance_score}%"],
            ['Total Requirements', str(report.total_requirements)],
            ['Compliant Requirements', str(report.compliant_requirements)],
            ['Violations Found', str(report.violation_count)]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Violations section
        if report.violations:
            story.append(Paragraph("Compliance Violations", styles['Heading2']))
            
            for violation in report.violations[:10]:  # Limit to first 10 violations
                story.append(Paragraph(f"<b>{violation.title}</b>", styles['Heading3']))
                story.append(Paragraph(f"Severity: {violation.severity.title()}", styles['Normal']))
                story.append(Paragraph(f"Description: {violation.description}", styles['Normal']))
                story.append(Paragraph(f"Detected: {violation.detected_at.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
                story.append(Spacer(1, 10))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.read()
    
    async def _export_html(self, report: ComplianceReport) -> str:
        """Export report as HTML"""
        template = Template('''
        <!DOCTYPE html>
        <html>
        <head>
            <title>{{ framework }} Compliance Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; }
                .header { border-bottom: 2px solid #333; padding-bottom: 20px; }
                .score { font-size: 2em; font-weight: bold; color: {{ score_color }}; }
                .metric { margin: 10px 0; }
                .violation { border-left: 4px solid #red; padding: 10px; margin: 10px 0; background: #f9f9f9; }
                .compliant { color: green; }
                .non-compliant { color: red; }
                table { border-collapse: collapse; width: 100%; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{{ framework }} Compliance Report</h1>
                <p>Generated: {{ generated_at }}</p>
                <p>Period: {{ period_start }} to {{ period_end }}</p>
            </div>
            
            <div class="executive-summary">
                <h2>Executive Summary</h2>
                <div class="score">Compliance Score: {{ compliance_score }}%</div>
                <div class="metric">Total Requirements: {{ total_requirements }}</div>
                <div class="metric">Compliant Requirements: {{ compliant_requirements }}</div>
                <div class="metric">Violations: {{ violation_count }}</div>
            </div>
            
            <div class="requirements">
                <h2>Requirements Assessment</h2>
                <table>
                    <tr><th>Requirement</th><th>Status</th><th>Confidence</th></tr>
                    {% for req in requirements_status %}
                    <tr>
                        <td>{{ req.requirement_id }}</td>
                        <td class="{{ 'compliant' if req.compliant else 'non-compliant' }}">
                            {{ 'Compliant' if req.compliant else 'Non-Compliant' }}
                        </td>
                        <td>{{ req.confidence_score }}%</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            
            {% if violations %}
            <div class="violations">
                <h2>Violations</h2>
                {% for violation in violations %}
                <div class="violation">
                    <h3>{{ violation.title }}</h3>
                    <p><strong>Severity:</strong> {{ violation.severity }}</p>
                    <p><strong>Description:</strong> {{ violation.description }}</p>
                    <p><strong>Detected:</strong> {{ violation.detected_at }}</p>
                </div>
                {% endfor %}
            </div>
            {% endif %}
        </body>
        </html>
        ''')
        
        score_color = '#28a745' if report.compliance_score >= 80 else '#ffc107' if report.compliance_score >= 60 else '#dc3545'
        
        return template.render(
            framework=report.framework.value.upper(),
            generated_at=report.generated_at.strftime('%Y-%m-%d %H:%M:%S'),
            period_start=report.report_period_start.date(),
            period_end=report.report_period_end.date(),
            compliance_score=report.compliance_score,
            total_requirements=report.total_requirements,
            compliant_requirements=report.compliant_requirements,
            violation_count=report.violation_count,
            requirements_status=report.requirements_status,
            violations=report.violations,
            score_color=score_color
        )
    
    async def _export_xml(self, report: ComplianceReport) -> str:
        """Export report as XML"""
        from xml.etree.ElementTree import Element, SubElement, tostring
        from xml.dom import minidom
        
        root = Element('ComplianceReport')
        root.set('framework', report.framework.value)
        root.set('generated_at', report.generated_at.isoformat())
        
        # Executive Summary
        exec_summary = SubElement(root, 'ExecutiveSummary')
        SubElement(exec_summary, 'ComplianceScore').text = str(report.compliance_score)
        SubElement(exec_summary, 'TotalRequirements').text = str(report.total_requirements)
        SubElement(exec_summary, 'CompliantRequirements').text = str(report.compliant_requirements)
        SubElement(exec_summary, 'ViolationCount').text = str(report.violation_count)
        
        # Requirements
        requirements = SubElement(root, 'Requirements')
        for req_status in report.requirements_status:
            req_elem = SubElement(requirements, 'Requirement')
            req_elem.set('id', req_status['requirement_id'])
            SubElement(req_elem, 'Title').text = req_status['title']
            SubElement(req_elem, 'Compliant').text = str(req_status['compliant']).lower()
            SubElement(req_elem, 'ConfidenceScore').text = str(req_status['confidence_score'])
        
        # Violations
        if report.violations:
            violations = SubElement(root, 'Violations')
            for violation in report.violations:
                violation_elem = SubElement(violations, 'Violation')
                violation_elem.set('id', violation.violation_id)
                SubElement(violation_elem, 'Severity').text = violation.severity
                SubElement(violation_elem, 'Title').text = violation.title
                SubElement(violation_elem, 'Description').text = violation.description
                SubElement(violation_elem, 'DetectedAt').text = violation.detected_at.isoformat()
                SubElement(violation_elem, 'Status').text = violation.status
        
        # Pretty print
        rough_string = tostring(root, 'unicode')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")
    
    async def _export_xlsx(self, report: ComplianceReport) -> bytes:
        """Export report as Excel XLSX"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ValueError("openpyxl library required for XLSX export")
        
        wb = openpyxl.Workbook()
        
        # Executive Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        ws_summary['A1'] = f"{report.framework.value.upper()} Compliance Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = "Generated At"
        ws_summary['B3'] = report.generated_at.strftime('%Y-%m-%d %H:%M:%S')
        
        ws_summary['A4'] = "Assessment Period"
        ws_summary['B4'] = f"{report.report_period_start.date()} to {report.report_period_end.date()}"
        
        ws_summary['A5'] = "Compliance Score"
        ws_summary['B5'] = f"{report.compliance_score}%"
        
        ws_summary['A6'] = "Total Requirements"
        ws_summary['B6'] = report.total_requirements
        
        ws_summary['A7'] = "Compliant Requirements"
        ws_summary['B7'] = report.compliant_requirements
        
        ws_summary['A8'] = "Violations"
        ws_summary['B8'] = report.violation_count
        
        # Requirements sheet
        ws_req = wb.create_sheet("Requirements")
        ws_req['A1'] = "Requirement ID"
        ws_req['B1'] = "Title"
        ws_req['C1'] = "Compliant"
        ws_req['D1'] = "Confidence Score"
        
        # Style header
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_req[col].font = Font(bold=True)
            ws_req[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add requirement data
        for idx, req_status in enumerate(report.requirements_status, 2):
            ws_req[f'A{idx}'] = req_status['requirement_id']
            ws_req[f'B{idx}'] = req_status['title']
            ws_req[f'C{idx}'] = 'Yes' if req_status['compliant'] else 'No'
            ws_req[f'D{idx}'] = f"{req_status['confidence_score']}%"
        
        # Violations sheet
        if report.violations:
            ws_viol = wb.create_sheet("Violations")
            ws_viol['A1'] = "Violation ID"
            ws_viol['B1'] = "Severity"
            ws_viol['C1'] = "Title"
            ws_viol['D1'] = "Description"
            ws_viol['E1'] = "Detected At"
            
            # Style header
            for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
                ws_viol[col].font = Font(bold=True)
                ws_viol[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add violation data
            for idx, violation in enumerate(report.violations, 2):
                ws_viol[f'A{idx}'] = violation.violation_id
                ws_viol[f'B{idx}'] = violation.severity
                ws_viol[f'C{idx}'] = violation.title
                ws_viol[f'D{idx}'] = violation.description
                ws_viol[f'E{idx}'] = violation.detected_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    
    # Additional helper methods for analysis...
    async def _generate_audit_summary(self, framework: ComplianceFramework, start_date: datetime, end_date: datetime, db: AsyncSession) -> Dict[str, Any]:
        """Generate audit log summary for the report period"""
        # Implementation would query audit logs and generate summary statistics
        return {
            'total_events': 0,
            'security_incidents': 0,
            'failed_authentications': 0,
            'administrative_actions': 0
        }
    
    async def _generate_processing_activities(self, framework: ComplianceFramework, start_date: datetime, end_date: datetime, db: AsyncSession) -> List[Dict[str, Any]]:
        """Generate data processing activities summary"""
        # Implementation would analyze data processing activities
        return []
    
    async def _perform_risk_assessment(self, violations: List[ComplianceViolation], requirements_status: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Perform risk assessment based on violations and compliance status"""
        critical_violations = len([v for v in violations if v.severity == 'critical'])
        high_violations = len([v for v in violations if v.severity == 'high'])
        
        if critical_violations > 0:
            risk_level = 'critical'
        elif high_violations > 3:
            risk_level = 'high'
        elif len(violations) > 5:
            risk_level = 'medium'
        else:
            risk_level = 'low'
        
        return {
            'overall_risk_level': risk_level,
            'key_findings': [
                f"Found {len(violations)} compliance violations",
                f"Critical violations: {critical_violations}",
                f"High severity violations: {high_violations}"
            ]
        }
    
    async def _generate_recommendations(self, violations: List[ComplianceViolation], requirements_status: List[Dict[str, Any]], compliance_score: float) -> List[str]:
        """Generate compliance recommendations"""
        recommendations = []
        
        if compliance_score < 80:
            recommendations.append("Implement comprehensive compliance monitoring program")
        
        if violations:
            recommendations.append(f"Address {len(violations)} identified compliance violations")
        
        critical_violations = [v for v in violations if v.severity == 'critical']
        if critical_violations:
            recommendations.append("URGENT: Address critical compliance violations immediately")
        
        return recommendations


# Global compliance reporting service instance
compliance_reporting_service = ComplianceReportingService()